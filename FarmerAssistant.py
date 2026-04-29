import streamlit as st
import requests
import pandas as pd
from bs4 import BeautifulSoup
import google.generativeai as genai
from deep_translator import GoogleTranslator
from PIL import Image
import io
from io import BytesIO
import re
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import base64
# ----------------------------
# UI Helpers & Styling
# ----------------------------
def apply_global_styles():
    # Subtle gradient background, card-like containers, and nicer headers
    st.markdown(
        """
        <style>
            /* Theme-aware background */
            :root[data-theme="light"] .stApp { background: linear-gradient(135deg, #f7fafc 0%, #eef2f7 100%); }
            :root[data-theme="dark"] .stApp { background: linear-gradient(135deg, #0f172a 0%, #0b1220 100%); }
            :root { --accent-color: #16a34a; --text-color: inherit; }
            .app-header {
                display: flex;
                align-items: center;
                gap: 12px;
                padding: 8px 0 16px 0;
                border-bottom: 1px solid rgba(0,0,0,0.06);
                margin-bottom: 12px;
            }
            .app-header h1 {
                font-size: 1.6rem;
                font-weight: 700;
                margin: 0;
                color: var(--text-color);
            }
            /* Hero banner with gradient overlay */
            .hero {
                position: relative;
                width: 100%;
                min-height: 180px;
                border-radius: 14px;
                overflow: hidden;
                margin: 4px 0 16px 0;
                border: 1px solid rgba(127,127,127,0.18);
            }
            .hero::before {
                content: "";
                position: absolute;
                inset: 0;
                background: var(--hero-gradient, linear-gradient(135deg, rgba(2,48,71,0.55), rgba(22,163,74,0.55)));
            }
            .hero .hero-bg {
                position: absolute; inset: 0; background-size: cover; background-position: center; filter: brightness(0.9);
            }
            :root[data-theme="dark"] .hero .hero-bg { filter: brightness(0.6); }
            .hero-content { position: relative; z-index: 1; padding: 18px; color: #fff; }
            .hero-title { font-size: 1.8rem; font-weight: 800; margin: 0 0 4px 0; }
            .hero-subtitle { opacity: 0.95; margin: 0; }
            /* Cards adapt to theme */
            :root[data-theme="light"] .nice-card { background: #ffffffcc; border: 1px solid rgba(0,0,0,0.06); box-shadow: 0 4px 14px rgba(0,0,0,0.06); }
            :root[data-theme="dark"] .nice-card { background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.12); box-shadow: 0 6px 18px rgba(0,0,0,0.35); }
            .nice-card { border-radius: 12px; padding: 16px 18px; margin: 10px 0 18px 0; color: var(--text-color); }
            /* Tighter metric alignment */
            [data-testid="stMetric"] {
                background: transparent;
                border: 1px solid rgba(127,127,127,0.2);
                border-radius: 10px;
                padding: 8px 10px;
            }
            /* Better buttons */
            .stButton button {
                border-radius: 10px;
                padding: 0.5rem 1rem;
                border: 1px solid rgba(127,127,127,0.3);
                transition: transform 0.06s ease;
            }
            .stButton button:hover { transform: translateY(-1px); box-shadow: 0 6px 20px rgba(0,0,0,0.08); }
            .stButton button:focus { outline: 2px solid var(--accent-color); }
            .st-emotion-cache-1jicfl2, .st-emotion-cache-1r6slb0 { color: inherit; }
            a, .stLinkButton button { color: var(--accent-color) !important; }
            /* Tables readable on both themes */
            :root[data-theme="dark"] table { color: var(--text-color); }
            :root[data-theme="dark"] thead th { background: rgba(255,255,255,0.06); }
            :root[data-theme="light"] thead th { background: rgba(0,0,0,0.03); }
        </style>
        """,
        unsafe_allow_html=True,
    )

def render_page_header(title_text, subtitle_text=None):
    lang = st.session_state.get("lang", "en")
    title_t = translate_text(title_text, lang)
    subtitle_t = translate_text(subtitle_text, lang) if subtitle_text else ""

    # Select contextual image per page
    images = {
        "Registration": "https://images.unsplash.com/photo-1500382017468-9049fed747ef?q=80&w=1600&auto=format&fit=crop",
        "Crop Recommendation": "https://images.unsplash.com/photo-1523419409543-7177f6b9c43a?q=80&w=1600&auto=format&fit=crop",
        "Weather": "https://images.unsplash.com/photo-1504384308090-c894fdcc538d?q=80&w=1600&auto=format&fit=crop",
        "Price Comparison": "https://images.unsplash.com/photo-1515165562835-c3b8f8875a2d?q=80&w=1600&auto=format&fit=crop",
        "Chatbot": "https://images.unsplash.com/photo-1518770660439-4636190af475?q=80&w=1600&auto=format&fit=crop",
        "Crop Disease Analysis": "https://images.unsplash.com/photo-1501004318641-b39e6451bec6?q=80&w=1600&auto=format&fit=crop",
        "Government Schemes": "https://images.unsplash.com/photo-1473186505569-9c61870c11f9?q=80&w=1600&auto=format&fit=crop",
        "Loan": "https://images.unsplash.com/photo-1454165804606-c3d57bc86b40?q=80&w=1600&auto=format&fit=crop"
    }
    # Infer key from title
    key = "Registration"
    if "Crop Recommendation" in title_text:
        key = "Crop Recommendation"
    elif "Weather" in title_text:
        key = "Weather"
    elif "Price" in title_text:
        key = "Price Comparison"
    elif "Chatbot" in title_text:
        key = "Chatbot"
    elif "Disease" in title_text:
        key = "Crop Disease Analysis"
    elif "Schemes" in title_text:
        key = "Government Schemes"
    elif "Loan" in title_text:
        key = "Loan"
    bg = images.get(key)
    # Top-left official logo using Streamlit API
    try:
        st.logo("logo3.png", size="large")
    except Exception:
        st.image("logo3.png", use_container_width=False)
    # Hero banner with gradient + background image and brand text
    st.markdown(
        f"""
        <div class='hero' style="--hero-gradient: linear-gradient(120deg, rgba(2,48,71,0.65), rgba(22,163,74,0.65));">
            <div class='hero-bg' style="background-image: url('{bg}');"></div>
            <div class='hero-content'>
                <div style='display:flex; gap:14px; align-items:center;'>
                    <img src="data:image/png;base64,{base64.b64encode(open('logo3.png','rb').read()).decode('utf-8')}" alt='logo' style='height:80px; width:auto; border-radius:10px; background: rgba(255,255,255,0.9); padding:6px;'>
                    <div style='display:flex; flex-direction:column;'>
                        <div class='hero-title'>{title_t}</div>
                        {f"<div class='hero-subtitle'>{subtitle_t}</div>" if subtitle_t else ''}
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def show_sidebar_branding():
    with st.sidebar:
        try:
            st.logo("logo3.png", size="large")
        except Exception:
            st.image("logo3.png", use_container_width=True)
        st.markdown("<div style='font-weight:800; font-size:1rem; margin-top:6px;'>Kishan AI</div>", unsafe_allow_html=True)
# ----------------------------
# API Keys & Config
# ----------------------------
GEMINI_API_KEY = "AIzaSyCKClnM1hYYJnm-6rq7eQKZ2Xaekga5tmA"
WEATHER_API_KEY = "c21c8ff5261c4f33b2572039251509"
ECOMMERCE_URLS = [
    "https://www.bighaat.com/collections/recommended-products-for-you",
    "https://agri-route.com/",
    "https://agripari.com/"
]
GOV_SCHEME_URL = "https://agriwelfare.gov.in/en/Major"

genai.configure(api_key=GEMINI_API_KEY)


# ----------------------------
# Translation Helper                                               
# ----------------------------
def translate_text(text, target_language):
    try:
        if target_language == "en":
            return text
        return GoogleTranslator(source="auto", target=target_language).translate(text)
    except Exception:
        return text


# ----------------------------
# Registration Form
# ----------------------------
def registration_page():
    render_page_header("Farmer Registration Form", "Create your profile to personalize all insights.")
    name = st.text_input(translate_text("Name of Farmer", st.session_state.lang))
    location = st.text_input(translate_text("Location", st.session_state.lang))
    plot_size = st.number_input(
        translate_text("Plot Size (in acres)", st.session_state.lang), min_value=0.0
    )
    soil_type = st.selectbox(
        translate_text("Soil Type", st.session_state.lang),
        ["Alluvial", "Black", "Red", "Laterite", "Arid"],
    )
    irrigation = st.selectbox(
        translate_text("Irrigation Method", st.session_state.lang),
        ["Canal", "Tube well", "Rainfed", "Drip", "Sprinkler"],
    )
    current_crop = st.text_input(
        translate_text("Current Crop", st.session_state.lang)
    )

    if st.button(translate_text("Submit", st.session_state.lang)):
        st.session_state.registration_data = {
            "name": name,
            "location": location,
            "plot_size": plot_size,
            "soil_type": soil_type,
            "irrigation": irrigation,
            "current_crop": current_crop,
        }
        st.success(translate_text("Registration Successful!", st.session_state.lang))


# ----------------------------
# Weather API (Detailed)
# ----------------------------
def get_weather_details(location):
    url = f"http://api.weatherapi.com/v1/current.json?key={WEATHER_API_KEY}&q={location}&aqi=yes"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            return response.json()
        else:

            return None
    except:
        return None


def weather_page():
    render_page_header("🌤️ Current Weather", "Live conditions and air quality for your farm.")

    location = st.text_input(
        translate_text("Enter city or location:", st.session_state.lang),
        st.session_state.registration_data["location"]
        if "registration_data" in st.session_state
        else "Delhi",
    )

    if st.button(translate_text("Get Weather", st.session_state.lang)):
        data = get_weather_details(location)

        if data:
            current = data["current"]
            location_info = data["location"]

            st.subheader(f"{location_info['name']}, {location_info['region']}, {location_info['country']}")
            st.write(f"🕒 Last updated: {current['last_updated']}")

            # Big Temperature
            st.markdown(f"## 🌡️ {current['temp_c']} °C ({current['temp_f']} °F) — {current['condition']['text']}")

            # Main metrics in columns
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Feels Like", f"{current['feelslike_c']} °C")
                st.metric("Wind Speed", f"{current['wind_kph']} kph")
                st.metric("Wind Direction", current['wind_dir'])
            with col2:
                st.metric("Humidity", f"{current['humidity']} %")
                st.metric("Pressure", f"{current['pressure_mb']} mb")
                st.metric("Cloud", f"{current['cloud']} %")
            with col3:
                st.metric("Visibility", f"{current['vis_km']} km")
                st.metric("UV Index", current['uv'])
                st.metric("Precipitation", f"{current['precip_mm']} mm")

            # Air Quality (if available)
            if "air_quality" in current:
                st.subheader("🌍 Air Quality")
                aq = current["air_quality"]
                st.write(f"**CO:** {aq.get('co', 'N/A'):.2f} µg/m³")
                st.write(f"**NO2:** {aq.get('no2', 'N/A'):.2f} µg/m³")
                st.write(f"**O3:** {aq.get('o3', 'N/A'):.2f} µg/m³")
                st.write(f"**SO2:** {aq.get('so2', 'N/A'):.2f} µg/m³")
                st.write(f"**PM2.5:** {aq.get('pm2_5', 'N/A'):.2f} µg/m³")
                st.write(f"**PM10:** {aq.get('pm10', 'N/A'):.2f} µg/m³")

        else:
            st.error(translate_text("Could not fetch weather. Please try again.", st.session_state.lang))

# ----------------------------
# Load and Prepare Crop Dataset using openpyxl
# ----------------------------
@st.cache_data
def load_crop_data(filepath):
    """Loads and prepares the crop recommendation dataset from an XLSX file using openpyxl."""
    try:
        # Fetch the file from URL
        response = requests.get(filepath)
        response.raise_for_status()  # Raise an error if request failed

        # Open the Excel file using openpyxl from bytes
        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet = wb.active  # Use the first sheet

        # Convert sheet data to a list of rows
        data = list(sheet.values)
        headers = data[0]  # First row as header
        rows = data[1:]    # Remaining rows as data

        # Create a DataFrame
        df = pd.DataFrame(rows, columns=headers)

        # --- Process temperature column ---
        temp_split = df['Temperature Range (°C)'].astype(str).str.split('-', expand=True)
        df['Min Temp'] = pd.to_numeric(temp_split[0], errors='coerce')
        df['Max Temp'] = pd.to_numeric(temp_split[1], errors='coerce').fillna(df['Min Temp'])

        # Drop rows where temperature conversion failed
        df.dropna(subset=['Min Temp', 'Max Temp'], inplace=True)

        return df

    except requests.exceptions.RequestException:
        st.error("Could not fetch the Excel file from the provided URL.")
        return None
    except Exception as e:
        st.error(f"Error loading Excel file: {e}")
        return None

# GitHub raw link to Excel file
CROP_DATA_PATH = "https://raw.githubusercontent.com/Aditya2811956/AI-based-FARMER_ASSISTANT/main/farm_last.xlsx"
crop_df = load_crop_data(CROP_DATA_PATH)



# ----------------------------
# Crop Recommendation (CORRECTED)
# ----------------------------
def crop_recommendation_page():
    render_page_header("Crop Recommendation 🌱", "Best crops based on climate, soil and irrigation.")
    
    if crop_df is None:
        st.error(translate_text("Crop dataset could not be loaded. Please check the file path.", st.session_state.lang))
        return

    if "registration_data" not in st.session_state or not st.session_state.registration_data.get('location'):
        st.warning(translate_text("Please complete your registration first, including your location.", st.session_state.lang))
        return

    reg_data = st.session_state.registration_data
    location = reg_data["location"]
    farmer_soil = reg_data["soil_type"]
    farmer_irrigation = reg_data["irrigation"]

    # Add a button to get recommendations
    if st.button(translate_text("Get Crop Recommendations", st.session_state.lang)):
        with st.spinner(translate_text("Fetching live weather and analyzing data...", st.session_state.lang)):
            
            # MODIFICATION: Use the same function as the weather_page
            weather_data = get_weather_details(location)

            if weather_data is None:
                st.error(translate_text("Could not fetch live weather data. Please check the location and try again.", st.session_state.lang))
                return

            # MODIFICATION: Extract the temperature from the detailed data
            temp = weather_data["current"]["temp_c"]

            # --- Recommendation Logic (No changes here) ---
            suitable_temp_crops = crop_df[(crop_df['Min Temp'] <= temp) & (crop_df['Max Temp'] >= temp)]

            if suitable_temp_crops.empty:
                st.warning(translate_text("No crops found suitable for the current temperature.", st.session_state.lang))
                return

            suitable_soil_crops = suitable_temp_crops[suitable_temp_crops['Soil Type'].str.contains(farmer_soil, case=False, na=False)]
            
            if suitable_soil_crops.empty:
                st.warning(f"{translate_text('No suitable crops for', st.session_state.lang)} {farmer_soil} {translate_text('soil at the current temperature.', st.session_state.lang)}")
                return

            final_recommendations = suitable_soil_crops[suitable_soil_crops['Best Irrigation Method'].str.contains(farmer_irrigation, case=False, na=False)]

        # This block is now correctly placed inside the button's scope
        st.subheader(translate_text("Top 3 Crop Recommendations", st.session_state.lang))
        st.write(f"**{translate_text('Location', st.session_state.lang)}:** {location} | **{translate_text('Current Temperature', st.session_state.lang)}:** {temp}°C | **{translate_text('Soil Type', st.session_state.lang)}:** {farmer_soil}")
        
        if not final_recommendations.empty:
            top_3_crops = final_recommendations.head(3)
            for index, row in top_3_crops.iterrows():
                st.success(f"**{row['Crop']}** ({row['Category']})")
                st.write(f"   - **{translate_text('Grows well in states like', st.session_state.lang)}:** {row['Major Growing States']}")
                st.write(f"   - **{translate_text('Ideal Rainfall', st.session_state.lang)}:** {row['Rainfall (mm)']} mm")
        else:
            st.error(translate_text("No crops perfectly match all your conditions (temperature, soil, and irrigation).", st.session_state.lang))
            st.info(f"{translate_text('Consider these options that match your soil type, even if irrigation differs:', st.session_state.lang)}")
            
            fallback_crops = suitable_soil_crops.head(3)
            for index, row in fallback_crops.iterrows():
                st.success(f"**{row['Crop']}** ({row['Category']})")
                st.write(f"   - **{translate_text('Grows well in states like', st.session_state.lang)}:** {row['Major Growing States']}")
                st.write(f"   - **{translate_text('Ideal Rainfall', st.session_state.lang)}:** {row['Rainfall (mm)']} mm")
                st.write(f"   - **{translate_text('Note', st.session_state.lang)}:** {translate_text('Recommended irrigation is', st.session_state.lang)} {row['Best Irrigation Method']}")
# Add this new import at the top of your file
from webdriver_manager.chrome import ChromeDriverManager


# ----------------------------
# Chatbot Page
# ----------------------------
def chatbot_page():
    render_page_header("Chatbot 🤖", "Ask anything. Get clear, translated answers.")
    lang = st.session_state.lang
    user_input = st.text_area(translate_text("Ask your question:", lang))

    if st.button(translate_text("Get Answer", lang)):
        if user_input:
            # Translate question to English
            user_input_en = (
                GoogleTranslator(source=lang, target="en").translate(user_input)
                if lang != "en"
                else user_input
            )

            model = genai.GenerativeModel("gemini-2.5-flash")
            response = model.generate_content(user_input_en)
            answer_en = response.text

            # Translate answer back
            answer_translated = (
                GoogleTranslator(source="en", target=lang).translate(answer_en)
                if lang != "en"
                else answer_en
            )

            st.subheader(translate_text("Answer in your language:", lang))
            st.write(answer_translated)

            if lang != "en":
                st.subheader("Answer (English):")
                st.write(answer_en)


# ----------------------------
# Crop Image Upload
# ----------------------------
def crop_image_page():
    render_page_header("Crop Disease Analysis 🌱", "Upload a leaf image to detect issues and treatments.")
    st.write(translate_text("Upload an image of a diseased plant, and the AI will identify the issue and suggest solutions.", st.session_state.lang))

    uploaded_file = st.file_uploader(
        translate_text("Choose an image...", st.session_state.lang),
        type=["jpg", "png", "jpeg"],
    )

    if uploaded_file is not None:
        # Display the uploaded image
        st.image(uploaded_file, caption=translate_text("Uploaded Plant Image", st.session_state.lang),  use_container_width=True)
        
        # Add a button to trigger the analysis
        if st.button(translate_text("Analyze Image", st.session_state.lang)):
            try:
                # Prepare the image and a more detailed prompt for the model
                image_bytes = uploaded_file.getvalue()
                pil_image = Image.open(io.BytesIO(image_bytes))

                # This improved prompt asks for a structured, detailed response.
                prompt = """
                You are an expert plant pathologist. Analyze the provided image of a diseased plant and provide a detailed report. Structure your response using the following markdown format:

                **Disease Identification:**
                (Identify the most likely disease. Be specific, e.g., "Tomato Early Blight" not just "Blight".)

                **Detailed Analysis:**
                (Describe the symptoms visible in the image and explain why they point to this diagnosis. Mention the common causes of this disease, such as environmental factors or pathogens.)

                **Organic Solutions:**
                (Provide a numbered list of actionable, organic treatment methods. Explain each step clearly.)

                **Chemical Solutions:**
                (Provide a numbered list of chemical treatment options, including the common active ingredients in fungicides/pesticides to look for. Include a disclaimer about following label instructions.)

                **Prevention Tips:**
                (Provide a numbered list of preventative measures to avoid this disease in the future.)
                """

                model = genai.GenerativeModel("gemini-2.5-flash")
                
                with st.spinner(translate_text("AI is analyzing the image, please wait...", st.session_state.lang)):
                    response = model.generate_content([prompt, pil_image])
                
                answer_en = response.text
                lang = st.session_state.lang
                answer_translated = translate_text(answer_en, lang)
                
                st.subheader(translate_text("Analysis Result", st.session_state.lang))
                st.markdown(answer_translated) # Use st.markdown to render the formatting

                if lang != "en":
                    st.subheader("Original Result (English)")
                    st.markdown(answer_en)
            
            except Exception as e:
                st.error(f"{translate_text('An error occurred during analysis:', st.session_state.lang)} {e}")

# ----------------------------
# Load Crop Price Dataset using openpyxl
# ----------------------------
@st.cache_data
def load_crop_price_data(filepath):
    """Loads the crop price dataset from an XLSX file using openpyxl."""
    try:
        # Fetch the file from URL
        response = requests.get(filepath)
        response.raise_for_status()  # Raise error if fetch fails

        # Open the Excel file using openpyxl
        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet = wb.active  # Use first sheet

        # Convert sheet data to a list of rows
        data = list(sheet.values)
        headers = data[0]  # First row as header
        rows = data[1:]    # Remaining rows as data

        # Create a DataFrame
        df = pd.DataFrame(rows, columns=headers)

        # Clean column names
        df.columns = [c.strip() for c in df.columns]

        # Clean price column (remove ₹ and convert to numeric)
        if "Price (₹ per Quintal)" in df.columns:
            df["Price (₹ per Quintal)"] = (
                df["Price (₹ per Quintal)"]
                .astype(str)
                .str.replace("₹", "", regex=False)
                .str.replace(",", "", regex=False)
                .str.strip()
            )
            df["Price (₹ per Quintal)"] = pd.to_numeric(df["Price (₹ per Quintal)"], errors="coerce")

        # Drop empty rows
        df = df.dropna(subset=["Commodity"])

        return df

    except requests.exceptions.RequestException:
        st.error(f"Could not fetch the Excel file from the URL: {filepath}")
        return None
    except Exception as e:
        st.error(f"Error loading crop price file: {e}")
        return None

# GitHub raw link to Excel file
CROP_PRICE_DATA_PATH = "https://raw.githubusercontent.com/Aditya2811956/AI-based-FARMER_ASSISTANT/main/crop_prices(2).xlsx"
crop_price_df = load_crop_price_data(CROP_PRICE_DATA_PATH)



# ----------------------------
# Current Price Page (from Dataset)
# ----------------------------
def current_price_page():
    render_page_header("Current Price 💹", "Check Minimum Support Prices (MSP) and crop rates.")

    if crop_price_df is None or crop_price_df.empty:
        st.error(translate_text("Crop price dataset could not be loaded. Please check the file path.", st.session_state.lang))
        return

    crop_name = st.text_input(
        translate_text("Enter Crop Name (e.g., Wheat, Rice, Tomato)", st.session_state.lang)
    )

    if st.button(translate_text("Get Current Price", st.session_state.lang)):
        if not crop_name:
            st.error(translate_text("Please enter a crop name.", st.session_state.lang))
            return

        # Filter dataset by crop name
        df_filtered = crop_price_df[crop_price_df['Commodity'].str.contains(crop_name, case=False, na=False)]

        if not df_filtered.empty:
            st.success(f"{translate_text('Found', st.session_state.lang)} {len(df_filtered)} {translate_text('records for', st.session_state.lang)} {crop_name}")
            st.dataframe(df_filtered[["Commodity", "Category", "Price (₹ per Quintal)", "Price Type / Basis"]])
        else:
            st.error(translate_text("No price data found for the given crop.", st.session_state.lang))


# ----------------------------
# Load Government Scheme Dataset using openpyxl
# ----------------------------
@st.cache_data
def load_scheme_data(filepath):
    """Loads the government schemes dataset from an XLSX file using openpyxl."""
    try:
        # Fetch the file from the URL
        response = requests.get(filepath)
        response.raise_for_status()  # Raise error if fetch fails

        # Load workbook using openpyxl
        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet = wb.active  # Use first sheet

        # Convert sheet data to a list of rows
        data = list(sheet.values)
        headers = data[0]  # First row as header
        rows = data[1:]    # Remaining rows as data

        # Create DataFrame
        df = pd.DataFrame(rows, columns=headers)

        return df

    except requests.exceptions.RequestException:
        st.error(f"Could not fetch the Excel file from the URL: {filepath}")
        return None
    except Exception as e:
        st.error(f"Error loading scheme file: {e}")
        return None

# GitHub raw link to the scheme Excel file
SCHEME_DATA_PATH = "https://raw.githubusercontent.com/Aditya2811956/AI-based-FARMER_ASSISTANT/main/scheme.xlsx"
scheme_df = load_scheme_data(SCHEME_DATA_PATH)


# ----------------------------
# Government Schemes (Updated to read from file)
# ----------------------------
def govt_schemes_page():
    render_page_header("Government Schemes 📜", "Explore benefits and apply quickly.")
    
    if st.button(translate_text("Fetch Schemes", st.session_state.lang)):
        if scheme_df is not None and not scheme_df.empty:
            st.success(f"{translate_text('Successfully loaded', st.session_state.lang)} {len(scheme_df)} {translate_text('schemes from the file.', st.session_state.lang)}")
            
            # Display the DataFrame from the Excel file as a table
            st.table(scheme_df)
        else:
            st.error(translate_text("Could not load the scheme data. Please ensure the 'scheme.xlsx' file is in the correct location.", st.session_state.lang))

# ----------------------------
# NEW: Loan EMI Calculator Page (with Interest Type Choice)
# ----------------------------
def loan_calculator_page():
    render_page_header("Loan EMI Calculator 💰", "Plan repayments with simple or EMI calculations.")

    # Input fields for the loan details
    principal = st.number_input(
        translate_text("Loan Amount (Principal)", st.session_state.lang), 
        min_value=1000, 
        value=100000, 
        step=1000
    )
    
    annual_rate = st.number_input(
        translate_text("Annual Interest Rate (%)", st.session_state.lang), 
        min_value=0.1, 
        value=10.5, 
        step=0.1
    )
    
    tenure_years = st.number_input(
        translate_text("Loan Tenure (in Years)", st.session_state.lang), 
        min_value=1, 
        value=5, 
        step=1
    )

    # NEW: Add a radio button to select the interest calculation method
    interest_type = st.radio(
        translate_text("Select Calculation Method", st.session_state.lang),
        [
            translate_text("Simple Interest", st.session_state.lang), 
            translate_text("Compound Interest (Standard EMI)", st.session_state.lang)
        ]
    )

    if st.button(translate_text("Calculate", st.session_state.lang)):
        # --- Input Validation ---
        if not (principal > 0 and annual_rate > 0 and tenure_years > 0):
            st.error(translate_text("Please enter valid details for all fields.", st.session_state.lang))
            return # Stop execution if inputs are invalid

        total_months = tenure_years * 12

        # --- Calculation Logic ---
        # Perform calculation based on the user's choice
        if interest_type == translate_text("Simple Interest", st.session_state.lang):
            # Simple Interest Calculation
            total_interest = (principal * annual_rate * tenure_years) / 100
            total_payment = principal + total_interest
            emi = total_payment / total_months
        
        else: # Compound Interest (Standard EMI) Calculation
            monthly_rate = (annual_rate / 12) / 100
            
            # EMI formula: P * r * (1+r)^n / ((1+r)^n - 1)
            try:
                emi = (principal * monthly_rate * (1 + monthly_rate)**total_months) / ((1 + monthly_rate)**total_months - 1)
                total_payment = emi * total_months
                total_interest = total_payment - principal
            except ZeroDivisionError:
                st.error(translate_text("Calculation resulted in an error. Please check your inputs.", st.session_state.lang))
                return

        # --- Display Results ---
        st.subheader(translate_text("Loan Repayment Details", st.session_state.lang))
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric(
                label=translate_text("Monthly Payment (EMI)", st.session_state.lang),
                value=f"₹ {emi:,.2f}"
            )
        with col2:
            st.metric(
                label=translate_text("Total Interest Payable", st.session_state.lang),
                value=f"₹ {total_interest:,.2f}"
            )
        with col3:
            st.metric(
                label=translate_text("Total Amount Payable", st.session_state.lang),
                value=f"₹ {total_payment:,.2f}"
            )


def loan_assistance_page():
    render_page_header("Loan Assistance & Information 🏦", "Checklist, scheme summaries and official portal links.")

    # Check if the scheme data is loaded
    if scheme_df is None or scheme_df.empty:
        st.error(translate_text("Loan scheme data is currently unavailable. Please check the file path.", st.session_state.lang))
        return

    st.write(translate_text("This section helps you understand and prepare for applying for agricultural loans through government schemes.", st.session_state.lang))

    # --- Section 1: Interactive Document Checklist ---
    st.subheader(translate_text("Step 1: Prepare Your Documents", st.session_state.lang))
    st.write(translate_text("Most agricultural loans require the following documents. Please check them off as you gather them:", st.session_state.lang))

    docs = {
        "Aadhaar Card": False,
        "PAN Card": False,
        "Proof of Land Ownership (Khasra/Khatauni)": False,
        "Bank Account Statement (Last 6 months)": False,
        "Passport Size Photographs": False
    }

    for doc, _ in docs.items():
        st.checkbox(translate_text(doc, st.session_state.lang))

    # --- Section 2: Explore Loan Schemes ---
    st.subheader(translate_text("Step 2: Explore Available Loan Schemes", st.session_state.lang))
    st.write(translate_text("Here are some of the major schemes available. Click on one to see details.", st.session_state.lang))

    # Display each scheme from your Excel file in an expander
    for index, row in scheme_df.iterrows():
        with st.expander(f"{row['Scheme Title']}"):
            st.write(f"**{translate_text('Published/Updated On', st.session_state.lang)}:** {row['Publish Date']}")
            
            # Use Gemini to generate a simple summary of the scheme based on its name
            if st.button(f"{translate_text('Get a Quick Summary for', st.session_state.lang)} {row['Scheme Title']}", key=f"summary_{index}"):
                with st.spinner(translate_text("Generating summary...", st.session_state.lang)):
                    try:
                        model = genai.GenerativeModel("gemini-2.5-flash")
                        prompt = f"In 100 words, briefly explain the purpose and key benefits of the '{row['Scheme Title']}' scheme for an Indian farmer."
                        response = model.generate_content(prompt)
                        st.markdown(response.text)
                    except Exception as e:
                        st.error(f"Could not generate summary: {e}")
            
            # Provide a direct link to apply
            st.link_button(
                translate_text("Visit Official Scheme Page", st.session_state.lang),
                row['Link']
            )
    
    # --- Section 3: Direct Link to National Portal ---
    st.subheader(translate_text("Step 3: Apply on the Official Portal", st.session_state.lang))
    st.write(translate_text("The **Jan Samarth** portal is the national platform for all credit-linked government schemes. You can apply for many agricultural loans here.", st.session_state.lang))
    st.link_button(
        translate_text("Go to Jan Samarth Portal", st.session_state.lang),
        "https://www.jansamarth.in/agri-loan"
    )

# ----------------------------
# Main App
# ----------------------------
def main():
    st.set_page_config(page_title="Farmer Assistant", layout="wide", page_icon="logo3.png")
    apply_global_styles()
    show_sidebar_branding()
    if "lang" not in st.session_state:
        st.session_state.lang = "en"

    lang_choice = st.sidebar.selectbox(
        "🌐 Language",
        [
            ("English", "en"),
            ("Hindi", "hi"),
            ("Marathi", "mr"),
            ("Kannada", "kn"),
            ("Malayalam", "ml"),
        ],
        format_func=lambda x: x[0],
    )
    st.session_state.lang = lang_choice[1]

    menu = st.sidebar.radio(
        "📌 Navigation",
        [
            "Registration",
            "Crop Recommendation",
            "Weather Report",
            "Chatbot 🤖",
            "Crop Disease Analysis 🌱",
            "Current Price 💹",
            "Government Schemes 📜",
            "Loan Calculator 💰",
            "Loan Assistance 🏦",
        ],
    )

    if menu == "Registration":
        registration_page()
    elif menu == "Crop Recommendation":
        crop_recommendation_page()
    elif menu == "Weather Report":
        weather_page()
    elif menu == "Chatbot 🤖":
        chatbot_page()
    elif menu == "Crop Disease Analysis 🌱":
        crop_image_page()
    elif menu == "Current Price 💹":
        current_price_page()
    elif menu == "Government Schemes 📜":
        govt_schemes_page()
    elif menu == "Loan Calculator 💰":
        loan_calculator_page()
    elif menu == "Loan Assistance 🏦": 
        loan_assistance_page()    

if __name__ == "__main__":
    main()